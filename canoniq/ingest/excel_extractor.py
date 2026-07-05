"""
Excel extractor — extracts metric candidates from Excel workbooks.

Targets:
  - Named ranges -> metric name + any formula attached
  - Cells with formula starting with = -> calculated measure candidates
  - Row/column header detection -> dimension candidates
  - Sheet names -> business domain classification

No LLM involved. Pure openpyxl parsing.
Requires: pip install openpyxl
"""

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from canoniq.ingest.base import DocumentMetricCandidate, SourceType, TableSchema

logger = logging.getLogger(__name__)

# Excel aggregation function patterns -> SQL equivalents
_EXCEL_TO_SQL_AGG = {
    "SUM": "SUM",
    "COUNT": "COUNT",
    "COUNTA": "COUNT",
    "COUNTIF": "COUNT",       # approximate
    "COUNTIFS": "COUNT",
    "AVERAGE": "AVG",
    "AVERAGEIF": "AVG",
    "AVERAGEIFS": "AVG",
    "MIN": "MIN",
    "MAX": "MAX",
}

_FORMULA_AGG_RE = re.compile(
    r"\b(" + "|".join(_EXCEL_TO_SQL_AGG) + r")\s*\(",
    re.IGNORECASE,
)


def _detect_agg(formula: str) -> str | None:
    """Return the SQL aggregation name for an Excel formula, or None."""
    match = _FORMULA_AGG_RE.search(formula)
    if not match:
        return None
    return _EXCEL_TO_SQL_AGG.get(match.group(1).upper())


def _nearest_label(sheet: Any, cell: Any) -> str | None:
    """
    Find the nearest non-empty, non-formula text cell to the left of
    or above `cell` — used as a human-readable metric name.
    """
    # Check cell directly to the left
    if cell.column > 1:
        left = sheet.cell(row=cell.row, column=cell.column - 1)
        if left.value and not str(left.value).startswith("="):
            return str(left.value).strip()
    # Check cell directly above
    if cell.row > 1:
        above = sheet.cell(row=cell.row - 1, column=cell.column)
        if above.value and not str(above.value).startswith("="):
            return str(above.value).strip()
    return None


def _clean_name(raw: str) -> str:
    """Convert a label like 'Total Net Revenue (USD)' to 'total_net_revenue'."""
    clean = re.sub(r"[^a-zA-Z0-9\s_]", "", raw)
    return re.sub(r"\s+", "_", clean.strip()).lower()


def extract_from_excel(
    path: str,
    schemas: list[TableSchema],
) -> list[DocumentMetricCandidate]:
    """
    Extract metric candidates from an Excel workbook.

    Returns DocumentMetricCandidate list. Candidates from named ranges
    get SourceType.EXCEL_NAMED (higher trust); candidates from bare
    formula cells get SourceType.EXCEL_FORMULA (lower trust).
    """
    logger.info("Extracting from Excel: %s", path)
    wb = openpyxl.load_workbook(path, data_only=False)
    results: list[DocumentMetricCandidate] = []

    # Build a flat set of real column names for grounding
    real_columns: set[str] = set()
    for schema in schemas:
        for col in schema.columns:
            real_columns.add(col.name.lower())

    # -- Pass 1: named ranges (highest trust) --------------------------------
    # wb.defined_names is a dict-like DefinedNameDict (name -> DefinedName);
    # there is no `.definedName` attribute on it.
    for named_range in wb.defined_names.values():
        name = named_range.name
        # Skip internal Excel names
        if name.startswith("_") or name.startswith("Print"):
            continue
        results.append(
            DocumentMetricCandidate(
                raw_name=name,
                raw_definition=f"Named range: {name}",
                raw_filter=None,
                resolved_expression=None,
                resolved_table=None,
                source_type=SourceType.EXCEL_NAMED,
                source_file=path,
                source_page=None,
                source_section=named_range.attr_text,
                has_approval_signal=False,
                grounding_confidence=0.3,   # named but no column mapping yet
                ambiguous=True,
            )
        )

    # -- Pass 2: formula cells ------------------------------------------------
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                formula = str(cell.value)
                if not formula.startswith("="):
                    continue

                agg_fn = _detect_agg(formula)
                if not agg_fn:
                    continue

                label = _nearest_label(sheet, cell)
                raw_name = (
                    _clean_name(label)
                    if label
                    else f"{sheet.title}_{get_column_letter(cell.column)}{cell.row}"
                )

                # Try to find a column name referenced in the formula
                formula_tokens = re.findall(r"\b[a-z_]+\b", formula.lower())
                matched_cols = [t for t in formula_tokens if t in real_columns]

                if matched_cols:
                    resolved_expr = f"{agg_fn}({matched_cols[0]})"
                    confidence = len(matched_cols) / max(len(formula_tokens), 1)
                else:
                    resolved_expr = None
                    confidence = 0.1

                results.append(
                    DocumentMetricCandidate(
                        raw_name=raw_name,
                        raw_definition=f"Excel formula: {formula} (sheet: {sheet.title})",
                        raw_filter=None,
                        resolved_expression=resolved_expr,
                        resolved_table=None,
                        source_type=SourceType.EXCEL_FORMULA,
                        source_file=path,
                        source_page=None,
                        source_section=sheet.title,
                        has_approval_signal=False,
                        grounding_confidence=min(confidence, 1.0),
                        ambiguous=resolved_expr is None,
                    )
                )

    logger.info(
        "Excel extraction complete: %d candidates from %s", len(results), Path(path).name
    )
    return results
